#!/usr/bin/python3

import datetime
import time

import pandas as pd
import matplotlib.pyplot as plt
import yahoo_fin.stock_info as yf_si
from openpyxl import load_workbook

import app_db
import config

def export_to_excel(user_id):
    tickers = app_db.load_tickers(user_id)
    wb = load_workbook(f'{config.FILEPATHS["download"]}{config.SYSTEM_SLASH}template.xlsx')
    #print (f'{config.FILEPATHS["download"]}{config.SYSTEM_SLASH}template.xlsx')
    sht = wb['MyTickers']
    start_row = 3
    for ticker in tickers:
        #print(ticker['id'], ticker['name'])
        cell = sht.cell(row=start_row, column=1)
        cell.value = ticker['id']
        cell = sht.cell(row=start_row, column=2)
        cell.value = ticker['name']
        start_row += 1
    wb.save(f'{config.FILEPATHS["download"]}{config.SYSTEM_SLASH}{user_id}.xlsx')

def import_from_excel(user_id, filename):
    dfs = pd.read_excel(filename, usecols=[0,1], skiprows=1)
    for ticker in zip(dfs['Тикер'], dfs['Название']):
        if not exist_ticker(ticker[0], user_id):
            try:
                app_db.add_ticker(ticker[0], ticker[1], user_id)
            except:
                app_db.write_to_log(user_id, f'Ошибка загрузки тикера {ticker[0]} из файла {filename} пользователем {user_id}')

def get_tickers_list():
    return {t[0]: t[1] for t in zip(yf_si.tickers_sp500(include_company_data=True)['Symbol'],
                                 yf_si.tickers_sp500(include_company_data=True)['Security'])}


def get_stock_data(ticker):
    now = datetime.datetime.now()
    longtime = now - datetime.timedelta(days=150)

    data = None
    ticker_data = None

    try:
        data = yf_si.get_data(ticker, longtime.strftime("%Y-%m-%d"), now.strftime("%Y-%m-%d"))
    except:
        pass

    if data is not None:
        ticker_data = [[index, row['adjclose'], 0, 0] for index, row in data.iterrows()]

        for i in range(20, len(ticker_data)):
            my_sum = 0
            for cur in range(i - 19, i + 1):
                my_sum += ticker_data[cur][1]
            ticker_data[i][3] = my_sum / 20

            my_sum = 0
            for cur in range(i - 4, i + 1):
                my_sum += ticker_data[cur][1]
            ticker_data[i][2] = my_sum / 5

    return ticker_data


def chart_to_image(ticker):
    ticker_data = get_stock_data(ticker)

    plt.style.use(['dark_background'])
    plt.plot([a[0] for a in ticker_data[-65:]], [a[1] for a in ticker_data[-65:]], label=f'{ticker}')
    plt.plot([a[0] for a in ticker_data[-65:]], [a[2] for a in ticker_data[-65:]], label='средняя за неделю')
    plt.plot([a[0] for a in ticker_data[-65:]], [a[3] for a in ticker_data[-65:]], label='средняя за месяц')
    plt.tick_params(axis='both', which='major', labelsize=8)
    plt.legend()
    plt.savefig(config.FILEPATHS['charts'] + config.SYSTEM_SLASH + f'{ticker}.png')
    plt.clf()
    plt.cla()
    plt.close()

    if len(ticker_data) < 60:
        return False
    else:
        return True


def ticker_price_to_ma(ticker):
    ticker_data = get_stock_data(ticker)
    if ticker_data is None:
        return "error", "error"
    now = len(ticker_data) - 1

    price_now = ticker_data[now][1]
    ma_20 = ticker_data[now][3]
    ma_5 = ticker_data[now][2]

    try:
        return round(price_now / ma_20 * 100 - 100), \
               round(price_now / ma_5 * 100 - 100)
    except:
        return "error", "error"


def ticker_price_to_ma_from_db(user_id, ticker):
    all_tickers = app_db.load_tickers(user_id)
    deviat_month = "error"
    deviat_week = "error"
    for tick in all_tickers:
        if ticker == tick['id']:
            deviat_month = tick['deviat_month']
            deviat_week = tick['deviat_week']
    return deviat_month, deviat_week


def ticker_ma_to_db():
    all_db_tickers = app_db.get_all_tickers()
    result = 0
    for ticker in all_db_tickers:
        ticker_ma20, ticker_ma5 = ticker_price_to_ma(ticker)
        if ticker_ma20 != 'error' and ticker_ma20 != 'error':
            app_db.set_ticker_ma(ticker, ticker_ma20, ticker_ma5)
            result += 1
        else:
            app_db.set_ticker_ma(ticker, 0, 0)
    return result


def ticker_ma_to_db_daemon():
    from api import tickers_cache
    while True:
        date = datetime.datetime.today()
        if date.strftime('%H') == '10':
            app_db.tickers_list_update()
            ticker_ma_to_db()
            app_db.write_to_log('default', f'Пересчитаны отклонения в базе')
            all_users = app_db.get_all_users()
            for user in all_users:
                tickers_cache.remove(user)
        time.sleep(3600)


def analize_tickers(user_id, limit=1):
    all_tickers = app_db.load_tickers(user_id)

    all_ticker_names = {}
    all_ticker_price_to_ma20 = {}
    all_ticker_price_to_ma5 = {}

    for ticker in all_tickers:
        tick_id = ticker['id']
        all_ticker_names[tick_id] = ticker['name']
        ticker_data20, ticker_data5 = ticker_price_to_ma_from_db(user_id, tick_id)

        if ticker_data20 != 'error' and ticker_data5 != 'error':
            if abs(ticker_data20) > limit:
                if ticker_data20 > 0:
                    all_ticker_price_to_ma20[tick_id] = "+" + str(ticker_data20) + "% \U00002705"
                else:
                    all_ticker_price_to_ma20[tick_id] = str(ticker_data20) + "% \U00002757"

            if abs(ticker_data5) > limit:
                if ticker_data5 > 0:
                    all_ticker_price_to_ma5[tick_id] = "+" + str(ticker_data5) + "% \U00002705"
                    if ticker_data5 > ticker_data20:
                        all_ticker_price_to_ma5[tick_id] += "\U00002705"
                else:
                    all_ticker_price_to_ma5[tick_id] = str(ticker_data5) + "% \U00002757"
                    if ticker_data5 < ticker_data20:
                        all_ticker_price_to_ma5[tick_id] += "\U00002757"

    out_message = ''

    if len(all_ticker_price_to_ma20) > 0:
        out_message += '\U0001F4C8 Рост/падение к среднемесячному значению:\n'
        for ticker in all_ticker_price_to_ma20:
            out_message += str(all_ticker_names[ticker]) + " " + str(all_ticker_price_to_ma20[ticker]) + " \n"

    if len(all_ticker_price_to_ma5) > 0:
        out_message += '\n\U0001F4C8 Рост/падение к средней за неделю:\n'
        for ticker in all_ticker_price_to_ma5:
            out_message += str(all_ticker_names[ticker]) + " " + str(all_ticker_price_to_ma5[ticker]) + " \n"

    if len(all_ticker_price_to_ma20) == 0 and len(all_ticker_price_to_ma5) == 0:
        out_message = '\U0001F4C8 Нет особых отклонений, рынок ровно дышит'

    return out_message


def probe_ticker(ticker):
    ticker_data = get_stock_data(ticker)
    return False if ticker_data is None else True


def exist_ticker(ticker, user_id):
    tickers = app_db.load_tickers(user_id)
    result = False
    for tick in tickers:
        if ticker == tick['id']:
            result = True
    return result

