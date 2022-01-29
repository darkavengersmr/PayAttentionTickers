#!/usr/bin/python3

import datetime
import asyncio

import pandas as pd
import matplotlib.pyplot as plt
import yahoo_fin.stock_info as yf_si
from openpyxl import load_workbook

import async_app_db
import config

async def export_to_excel(user_id):
    tickers = await async_app_db.load_tickers(user_id)
    wb = load_workbook(f'{config.FILEPATHS["download"]}{config.SYSTEM_SLASH}template.xlsx')
    sht = wb['MyTickers']
    start_row = 3
    for ticker in tickers:
        cell = sht.cell(row=start_row, column=1)
        cell.value = ticker['id']
        cell = sht.cell(row=start_row, column=2)
        cell.value = ticker['name']
        start_row += 1
    wb.save(f'{config.FILEPATHS["download"]}{config.SYSTEM_SLASH}{user_id}.xlsx')

async def import_from_excel(user_id, filename):
    dfs = pd.read_excel(filename, usecols=[0,1], skiprows=1)
    for ticker in zip(dfs['Тикер'], dfs['Название']):
        if not await exist_ticker(ticker[0], user_id):
            try:
                await async_app_db.add_ticker(ticker[0], ticker[1], user_id)
            except:
                await async_app_db.write_to_log(user_id, f'Ошибка загрузки тикера {ticker[0]} из файла {filename} пользователем {user_id}')

async def get_tickers_list():
    return {t[0]: t[1] for t in zip(yf_si.tickers_sp500(include_company_data=True)['Symbol'],
                                 yf_si.tickers_sp500(include_company_data=True)['Security'])}


async def get_stock_data(ticker):
    now = datetime.datetime.now()
    longtime = now - datetime.timedelta(days=150)

    data = None
    ticker_data = None

    try:
        data = yf_si.get_data(ticker, longtime.strftime("%Y-%m-%d"), now.strftime("%Y-%m-%d"))
    except:
        pass

    if data is not None:
        ticker_data = [[index, row['adjclose'], 0, 0] for index, row in data.iterrows()]

        for i in range(20, len(ticker_data)):
            my_sum = 0
            for cur in range(i - 19, i + 1):
                my_sum += ticker_data[cur][1]
            ticker_data[i][3] = my_sum / 20

            my_sum = 0
            for cur in range(i - 4, i + 1):
                my_sum += ticker_data[cur][1]
            ticker_data[i][2] = my_sum / 5

    return ticker_data


async def chart_to_image(ticker):
    ticker_data = await get_stock_data(ticker)

    if ticker_data is None:
        return False

    plt.style.use(['dark_background'])
    plt.plot([a[0] for a in ticker_data[-65:]], [a[1] for a in ticker_data[-65:]], label=f'{ticker}')
    plt.plot([a[0] for a in ticker_data[-65:]], [a[2] for a in ticker_data[-65:]], label='средняя за неделю')
    plt.plot([a[0] for a in ticker_data[-65:]], [a[3] for a in ticker_data[-65:]], label='средняя за месяц')
    plt.tick_params(axis='both', which='major', labelsize=8)
    plt.legend()
    plt.savefig(config.FILEPATHS['charts'] + config.SYSTEM_SLASH + f'{ticker}.png')
    plt.clf()
    plt.cla()
    plt.close()
    return True


async def ticker_price_to_ma(ticker):
    ticker_data = await get_stock_data(ticker)
    try:
        now = len(ticker_data) - 1

        price_now = ticker_data[now][1]
        ma_20 = ticker_data[now][3]
        ma_5 = ticker_data[now][2]

        return round(price_now / ma_20 * 100 - 100), \
               round(price_now / ma_5 * 100 - 100)
    except:
        return "error", "error"


async def ticker_price_to_ma_from_db(user_id, ticker):
    all_tickers = await async_app_db.load_tickers(user_id)
    deviat_month = "error"
    deviat_week = "error"
    for tick in all_tickers:
        if ticker == tick['id']:
            deviat_month = tick['deviat_month']
            deviat_week = tick['deviat_week']
    return deviat_month, deviat_week


async def ticker_ma_to_db():
    all_db_tickers = await async_app_db.get_all_tickers()
    result = 0
    for ticker in all_db_tickers:
        ticker_ma20, ticker_ma5 = await ticker_price_to_ma(ticker)
        if ticker_ma20 != 'error' and ticker_ma20 != 'error':
            await async_app_db.set_ticker_ma(ticker, ticker_ma20, ticker_ma5)
            result += 1
        else:
            await async_app_db.set_ticker_ma(ticker, 0, 0)
    return result


async def ticker_ma_to_db_daemon():
    from async_api import tickers_cache
    while True:
        date = datetime.datetime.today()
        if date.strftime('%H') == '10':
            await async_app_db.tickers_list_update()
            await ticker_ma_to_db()
            await async_app_db.write_to_log('default', f'Пересчитаны отклонения в базе')
            '''
            all_users = await async_app_db.get_all_users()
            for user in all_users:
                tickers_cache.delete(user)
            '''
            await tickers_cache.cache.flushdb()
        await asyncio.sleep(3600)


async def probe_ticker(ticker):
    ticker_data = await get_stock_data(ticker)
    return False if ticker_data is None else True


async def exist_ticker(ticker, user_id):
    tickers = await async_app_db.load_tickers(user_id)
    result = False
    for tick in tickers:
        if ticker == tick['id']:
            result = True
    return result

async def analize_tickers(user_id, limit=1):
    all_tickers = await async_app_db.load_tickers(user_id)

    full_tickers_data = []

    for ticker in all_tickers:
        ticker_data = {}
        ticker_data['id'] = ticker['id']
        ticker_data['name'] = ticker['name']

        ticker_data20, ticker_data5 = await ticker_price_to_ma_from_db(user_id, ticker_data['id'])

        if ticker_data20 != 'error' and ticker_data5 != 'error':
            ticker_data['deviat_month'] = ticker_data20
            ticker_data['deviat_week'] = ticker_data5

            if abs(ticker_data['deviat_month']) > limit:
                if ticker_data['deviat_month'] > 0:
                    ticker_data['show_month'] = f'{ticker_data["name"]} +{ticker_data["deviat_month"]}% \U00002705'
                else:
                    ticker_data['show_month'] = f'{ticker_data["name"]} {ticker_data["deviat_month"]}% \U00002757'

            if abs(ticker_data['deviat_week']) > limit:
                if ticker_data['deviat_week'] > 0:
                    ticker_data['show_week'] = f'{ticker_data["name"]} +{ticker_data["deviat_week"]}% \U00002705'
                    if ticker_data['deviat_week'] > ticker_data['deviat_month']:
                        ticker_data['show_week'] = f'{ticker_data["show_week"]}\U00002705'
                else:
                    ticker_data['show_week'] = f' {ticker_data["name"]} {ticker_data["deviat_week"]}% \U00002757'
                    if ticker_data['deviat_week'] < ticker_data['deviat_month']:
                        ticker_data['show_week'] = f'{ticker_data["show_week"]}\U00002757'

            full_tickers_data.append(ticker_data)

    out_message = '\U0001F4C8 Рост/падение к среднемесячному значению:\n'
    out_message_count = 0

    full_tickers_data.sort(key=lambda dictionary: dictionary['deviat_month'])

    for ticker in full_tickers_data:
        if 'show_month' in ticker:
            out_message += f'{ticker["show_month"]}\n'
            out_message_count += 1

    out_message += '\n\U0001F4C8 Рост/падение к средней за неделю:\n'

    full_tickers_data.sort(key=lambda dictionary: dictionary['deviat_week'])

    for ticker in full_tickers_data:
        if 'show_week' in ticker:
            out_message += f'{ticker["show_week"]}\n'
            out_message_count += 1

    if out_message_count == 0:
        out_message = '\U0001F4C8 Нет особых отклонений, рынок ровно дышит'

    return out_message